"""
Генератор Excel файлов (упрощенная версия)
"""
import logging
import os
from typing import Dict, List
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def generate_excel(application_data: Dict, output_dir: str = "/tmp") -> str:
    """
    Генерация Excel файла с упрощенной структурой
    
    Args:
        application_data: Данные заявки
        output_dir: Директория для сохранения файла
        
    Returns:
        str: Путь к сгенерированному файлу
    """
    logger.info("Начинаем генерацию Excel файла (упрощенная версия)")
    
    # Генерируем имя выходного файла
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    city = application_data.get('city', 'Unknown').replace('/', '_')
    output_filename = f"Заявка_{city}_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    # Создаем новую книгу
    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active
    ws.title = "Заявка на СМ"
    
    # Настройка стилей
    header_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=11)
    title_font = Font(name='Arial', size=14, bold=True)
    
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовок
    ws['A1'] = "ЗАЯВКА НА СЛУЖЕБНУЮ ПОЕЗДКУ"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')
    
    # Дата создания
    current_row = 2
    ws[f'A{current_row}'] = f"Дата подачи: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws[f'A{current_row}'].font = normal_font
    
    # Основная информация
    current_row += 2
    
    # Вид спорта
    ws[f'A{current_row}'] = "Вид спорта:"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'B{current_row}'] = application_data.get("sport_type", "")
    ws[f'B{current_row}'].font = normal_font
    ws.merge_cells(f'B{current_row}:F{current_row}')
    current_row += 1
    
    # Ранг мероприятия
    ws[f'A{current_row}'] = "Ранг мероприятия:"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'B{current_row}'] = application_data.get("event_rank", "")
    ws[f'B{current_row}'].font = normal_font
    ws.merge_cells(f'B{current_row}:F{current_row}')
    current_row += 1
    
    # Страна
    ws[f'A{current_row}'] = "Страна:"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'B{current_row}'] = application_data.get("country", "")
    ws[f'B{current_row}'].font = normal_font
    ws.merge_cells(f'B{current_row}:F{current_row}')
    current_row += 1
    
    # Город
    ws[f'A{current_row}'] = "Город:"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'B{current_row}'] = application_data.get("city", "")
    ws[f'B{current_row}'].font = normal_font
    ws.merge_cells(f'B{current_row}:F{current_row}')
    current_row += 2
    
    # Таблица участников
    ws[f'A{current_row}'] = "СПИСОК УЧАСТНИКОВ"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'A{current_row}:F{current_row}')
    current_row += 1
    
    # Заголовки таблицы
    headers = ['№', 'ФИО участника', 'Дата начала', 'Дата окончания']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    current_row += 1
    
    # Заполняем участников
    participants: List[Dict] = application_data.get("participants", [])
    
    for idx, participant in enumerate(participants, start=1):
        # Номер
        cell = ws.cell(row=current_row, column=1)
        cell.value = idx
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # ФИО
        cell = ws.cell(row=current_row, column=2)
        cell.value = participant.get("full_name", "")
        cell.font = normal_font
        cell.border = thin_border
        
        # Дата начала
        cell = ws.cell(row=current_row, column=3)
        cell.value = participant.get("date_from", "")
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # Дата окончания
        cell = ws.cell(row=current_row, column=4)
        cell.value = participant.get("date_to", "")
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        current_row += 1
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # Сохраняем файл
    wb.save(output_path)
    wb.close()
    
    logger.info(f"Excel файл успешно создан (упрощенная версия): {output_path}")
    return output_path
